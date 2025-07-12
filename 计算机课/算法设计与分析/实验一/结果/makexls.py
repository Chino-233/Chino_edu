import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 从提供的结果文件中提取数据
def extract_data_from_results():
    algorithms = ['quickSort', 'mergeSort', 'insertionSort', 'selectionSort', 'bubbleSort']
    
    # 存储所有算法的数据
    all_data = {}
    
    for algo in algorithms:
        filename = f"d:\\chino_edu\\makenum\\result\\{algo}_results.txt"
        try:
            # 首先尝试使用UTF-8编码
            with open(filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()
        except UnicodeDecodeError:
            try:
                # 如果UTF-8失败，尝试使用latin-1编码（它可以读取任何字节序列）
                with open(filename, 'r', encoding='latin-1') as file:
                    lines = file.readlines()
            except Exception as e:
                print(f"无法读取文件 {filename}: {e}")
                continue
            
        data = {}
        for line in lines:
            if line.strip() and not line.startswith("=") and not line.startswith("时间复杂度") and "-" not in line:
                if "数据规模" not in line:  # 跳过标题行
                    parts = line.strip().split("\t")
                    if len(parts) >= 2:
                        try:
                            size = int(parts[0].strip())
                            time = float(parts[1].strip())
                            data[size] = time
                        except ValueError:
                            print(f"无法解析行: {line}")
        
        all_data[algo] = data
    
    return all_data


def create_excel_tables(data):
    # 创建一个Excel工作簿
    wb = Workbook()
    
    # 创建第一个表 - 五种排序算法在10万到100万元素的性能比较
    ws1 = wb.active
    ws1.title = "排序算法比较(10万-100万)"
    
    # 设置表头
    ws1['A1'] = "算法"
    sizes = [100000, 250000, 500000, 750000, 1000000]
    for i, size in enumerate(sizes):
        col = get_column_letter(i + 2)
        ws1[f'{col}1'] = f"{size:,}元素"
    
    # 填充数据
    algorithms = ['quickSort', 'mergeSort', 'insertionSort', 'selectionSort', 'bubbleSort']
    algo_names = ["快速排序", "归并排序", "插入排序", "选择排序", "冒泡排序"]
    
    for i, (algo, name) in enumerate(zip(algorithms, algo_names)):
        ws1[f'A{i+2}'] = name
        for j, size in enumerate(sizes):
            col = get_column_letter(j + 2)
            if algo in data and size in data[algo]:
                value = data[algo][size]
                if value < 100:
                    ws1[f'{col}{i+2}'] = f"{value:.2f} ms"
                else:
                    ws1[f'{col}{i+2}'] = f"{int(value)} ms"
            else:
                ws1[f'{col}{i+2}'] = "N/A"
                print(f"警告：算法 {algo} 中没有大小为 {size} 的数据")
    
    # 创建第二个表 - 归并和快速排序在1000万元素的性能比较
    ws2 = wb.create_sheet("高效排序算法(1000万)")
    
    # 设置表头
    ws2['A1'] = "算法"
    ws2['B1'] = "10,000,000元素"
    
    # 填充数据
    high_perf_algos = ['quickSort', 'mergeSort']
    high_perf_names = ["快速排序", "归并排序"]
    
    for i, (algo, name) in enumerate(zip(high_perf_algos, high_perf_names)):
        ws2[f'A{i+2}'] = name
        size = 10000000
        if algo in data and size in data[algo]:
            value = data[algo][size]
            ws2[f'B{i+2}'] = f"{value:.2f} ms"
        else:
            ws2[f'B{i+2}'] = "N/A"
            print(f"警告：算法 {algo} 中没有大小为 {size} 的数据")
    
    # 设置格式
    for ws in [ws1, ws2]:
        # 设置列宽 - 修正列宽设置逻辑
        ws.column_dimensions['A'].width = 15
        for i in range(1, ws.max_column):
            col = get_column_letter(i + 1)
            ws.column_dimensions[col].width = 18
        
        # 设置表头格式
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 设置单元格边框
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # 数据行
                    cell.alignment = Alignment(horizontal='right')
    
    # 保存文件
    wb.save("排序算法性能比较.xlsx")
    print("Excel表格已生成: 排序算法性能比较.xlsx")

# 执行程序
algorithm_data = extract_data_from_results()
create_excel_tables(algorithm_data)